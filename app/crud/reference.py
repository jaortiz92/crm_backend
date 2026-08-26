# Python
from fastapi import HTTPException, status, UploadFile
from sqlalchemy.orm import Session
from typing import List
import openpyxl
from io import BytesIO
from openpyxl import Workbook
from decimal import Decimal

# App
from app.models.reference import Reference as ReferenceModel
from app.models.brand import Brand as BrandModel
from app.models.collection import Collection as CollectionModel
from app.models.budget.actualCost import ActualCost as ActualCostModel
from app.schemas.reference import (
    ReferenceCreate, Reference as ReferenceSchema,
    BulkUploadResult, BulkDeleteResult
)
from app.core import Gender


def get_references(db: Session, skip: int = 0, limit: int = 10) -> List[ReferenceSchema]:
    return db.query(ReferenceModel).offset(skip).limit(limit).all()


def get_reference(db: Session, id_reference: int) -> ReferenceSchema:
    result = db.query(ReferenceModel).filter(
        ReferenceModel.id_reference == id_reference
    ).first()
    return result


def create_reference(db: Session, reference: ReferenceCreate) -> ReferenceSchema:
    # Validate brand exists
    brand = db.query(BrandModel).filter(
        BrandModel.id_brand == reference.id_brand
    ).first()
    if not brand:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Brand with id {reference.id_brand} not found"
        )
    
    # Validate collection exists if provided
    if reference.id_collection:
        collection = db.query(CollectionModel).filter(
            CollectionModel.id_collection == reference.id_collection
        ).first()
        if not collection:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Collection with id {reference.id_collection} not found"
            )
    
    # Check if reference already exists
    existing = db.query(ReferenceModel).filter(
        ReferenceModel.reference == reference.reference
    ).first()
    if existing:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Reference '{reference.reference}' already exists"
        )
    
    try:
        data = reference.model_dump()
        if data.get('description'):
            data['description'] = data['description'].upper()
        db_reference = ReferenceModel(**data)
        db.add(db_reference)
        db.commit()
        db.refresh(db_reference)
        return db_reference
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creating reference: {str(e)}"
        )


def update_reference(db: Session, id_reference: int, reference: ReferenceCreate) -> ReferenceSchema:
    db_reference = db.query(ReferenceModel).filter(
        ReferenceModel.id_reference == id_reference
    ).first()
    
    if not db_reference:
        return None
    
    # Validate brand exists
    brand = db.query(BrandModel).filter(
        BrandModel.id_brand == reference.id_brand
    ).first()
    if not brand:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Brand with id {reference.id_brand} not found"
        )
    
    # Validate collection exists if provided
    if reference.id_collection:
        collection = db.query(CollectionModel).filter(
            CollectionModel.id_collection == reference.id_collection
        ).first()
        if not collection:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Collection with id {reference.id_collection} not found"
            )
    
    # Check if reference code already exists (excluding current record)
    if reference.reference != db_reference.reference:
        existing = db.query(ReferenceModel).filter(
            ReferenceModel.reference == reference.reference
        ).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Reference '{reference.reference}' already exists"
            )
    
    data = reference.model_dump()
    if data.get('description'):
        data['description'] = data['description'].upper()
    
    for key, value in data.items():
        setattr(db_reference, key, value)
    
    db.commit()
    db.refresh(db_reference)
    return db_reference


def delete_reference(db: Session, id_reference: int) -> bool:
    db_reference = db.query(ReferenceModel).filter(
        ReferenceModel.id_reference == id_reference
    ).first()
    
    if not db_reference:
        return False
    
    # Check if reference is used in actual_costs
    usage_count = db.query(ActualCostModel).filter(
        ActualCostModel.id_reference == id_reference
    ).count()
    
    if usage_count > 0:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Reference '{db_reference.reference}' cannot be deleted because it is used in {usage_count} actual cost records"
        )
    
    db.delete(db_reference)
    db.commit()
    return True


def bulk_upload_references(db: Session, file: UploadFile) -> BulkUploadResult:
    """
    Process Excel file for bulk upsert of references.
    Columns: Referencia, Marca, Descripción, Género, Valor Base, Colección
    """
    errors = []
    inserted = 0
    updated = 0
    
    try:
        # Read Excel file
        contents = file.file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Validate headers
        headers = [cell.value for cell in ws[1]]
        expected_headers = ["Referencia", "Marca", "Descripción", "Género", "Valor Base", "Colección"]
        
        if headers != expected_headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid headers. Expected: {expected_headers}, Got: {headers}"
            )
        
        # Process each row
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
            
            try:
                referencia = str(row[0]).strip() if row[0] else None
                marca = str(row[1]).strip() if row[1] else None
                descripcion = str(row[2]).strip().upper() if row[2] else None
                genero = str(row[3]).strip().upper() if row[3] else None
                valor_base = row[4] if row[4] is not None else None
                coleccion = str(row[5]).strip() if row[5] and str(row[5]).strip() else None
                
                # Validate required fields
                if not referencia:
                    errors.append(f"Row {row_num}: Referencia is required")
                    continue
                if not marca:
                    errors.append(f"Row {row_num}: Marca is required")
                    continue
                if not genero:
                    errors.append(f"Row {row_num}: Género is required")
                    continue
                if valor_base is None:
                    errors.append(f"Row {row_num}: Valor Base is required")
                    continue
                
                # Validate gender
                gender_map = {"U": Gender.U, "M": Gender.M, "F": Gender.F}
                if genero not in gender_map:
                    errors.append(f"Row {row_num}: Género must be U, M, or F")
                    continue
                
                # Validate valor_base is numeric
                try:
                    valor_base_float = float(valor_base)
                    if valor_base_float < 0:
                        errors.append(f"Row {row_num}: Valor Base must be >= 0")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Row {row_num}: Valor Base must be numeric")
                    continue
                
                # Search brand by name
                brand = db.query(BrandModel).filter(
                    BrandModel.brand_name == marca
                ).first()
                if not brand:
                    errors.append(f"Row {row_num}: Brand '{marca}' not found")
                    continue
                
                # Search collection by short name (if provided)
                collection_id = None
                if coleccion:
                    collection = db.query(CollectionModel).filter(
                        CollectionModel.short_collection_name == coleccion
                    ).first()
                    if not collection:
                        errors.append(f"Row {row_num}: Collection '{coleccion}' not found")
                        continue
                    collection_id = collection.id_collection
                
                # Check if reference already exists
                existing_ref = db.query(ReferenceModel).filter(
                    ReferenceModel.reference == referencia
                ).first()
                
                if existing_ref:
                    # Update existing reference
                    existing_ref.id_brand = brand.id_brand
                    existing_ref.description = descripcion
                    existing_ref.gender = gender_map[genero]
                    existing_ref.value_base = Decimal(str(valor_base_float))
                    existing_ref.id_collection = collection_id
                    updated += 1
                else:
                    # Insert new reference
                    new_ref = ReferenceModel(
                        reference=referencia,
                        id_brand=brand.id_brand,
                        description=descripcion,
                        gender=gender_map[genero],
                        value_base=Decimal(str(valor_base_float)),
                        id_collection=collection_id
                    )
                    db.add(new_ref)
                    inserted += 1
                    
            except Exception as e:
                errors.append(f"Row {row_num}: Error processing row - {str(e)}")
                continue
        
        # If there are errors, rollback and return them
        if errors:
            db.rollback()
            return BulkUploadResult(
                message="Upload failed with errors",
                total_filas=row_num - 1,
                insertadas=0,
                actualizadas=0,
                errores=errors
            )
        
        # Commit all changes
        db.commit()
        
        return BulkUploadResult(
            message="Upload completed successfully",
            total_filas=row_num - 1,
            insertadas=inserted,
            actualizadas=updated,
            errores=[]
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing file: {str(e)}"
        )


def bulk_delete_references(db: Session, file: UploadFile) -> BulkDeleteResult:
    """
    Process Excel file for bulk delete of references.
    Column: Referencia
    """
    try:
        # Read Excel file
        contents = file.file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Validate headers
        headers = [cell.value for cell in ws[1]]
        expected_headers = ["Referencia"]
        
        if headers != expected_headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid headers. Expected: {expected_headers}, Got: {headers}"
            )
        
        # Collect all references to delete
        references_to_delete = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
            
            referencia = str(row[0]).strip() if row[0] else None
            if referencia:
                references_to_delete.append(referencia)
        
        if not references_to_delete:
            return BulkDeleteResult(
                message="No references to delete",
                total_eliminadas=0,
                referencias_eliminadas=[]
            )
        
        # Validate NONE are in use in actual_costs
        references_in_use = []
        for referencia in references_to_delete:
            ref = db.query(ReferenceModel).filter(
                ReferenceModel.reference == referencia
            ).first()
            
            if not ref:
                continue  # Skip if doesn't exist
            
            usage_count = db.query(ActualCostModel).filter(
                ActualCostModel.id_reference == ref.id_reference
            ).count()
            
            if usage_count > 0:
                references_in_use.append(f"{referencia} (used in {usage_count} records)")
        
        # If any are in use, reject entire operation
        if references_in_use:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Cannot delete references. The following are in use: {', '.join(references_in_use)}"
            )
        
        # Delete all references
        deleted_refs = []
        for referencia in references_to_delete:
            ref = db.query(ReferenceModel).filter(
                ReferenceModel.reference == referencia
            ).first()
            
            if ref:
                db.delete(ref)
                deleted_refs.append(referencia)
        
        db.commit()
        
        return BulkDeleteResult(
            message="Bulk delete completed successfully",
            total_eliminadas=len(deleted_refs),
            referencias_eliminadas=deleted_refs
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing file: {str(e)}"
        )


def create_upload_template() -> bytes:
    """
    Generate Excel template for bulk upload.
    Columns: Referencia, Marca, Descripción, Género, Valor Base, Colección
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Referencias"
    
    # Add headers
    headers = ["Referencia", "Marca", "Descripción", "Género", "Valor Base", "Colección"]
    ws.append(headers)
    
    # Add example data in row 2
    ws.append(["REF001", "NIKE", "Zapatillas deportivas", "U", 150000.00, "INV25"])
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_delete_template() -> bytes:
    """
    Generate Excel template for bulk delete.
    Column: Referencia
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Referencias"
    
    # Add headers
    headers = ["Referencia"]
    ws.append(headers)
    
    # Add example data in row 2
    ws.append(["REF001"])
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
